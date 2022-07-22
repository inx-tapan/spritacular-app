import os
import sys
import urllib.request
import uuid
from io import BytesIO

import openpyxl
import requests
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.files.temp import NamedTemporaryFile
from django.core.files.uploadedfile import InMemoryUploadedFile

from PIL import Image


from openpyxl_image_loader import SheetImageLoader
from django.core.management.base import BaseCommand
from spritacular.settings import BASE_DIR
from quiz.models import Question, QuizOption


class Command(BaseCommand):
    help = 'This class is for populating the quiz questions data.'

    def handle(self, *args, **options):
        # try:
            print(os.path.join(BASE_DIR, 'labelled_imagery_updated.xlsx'))
            wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'labelled_imagery_updated.xlsx'))
            sheet = wb['Sheet1']
            # calling the image_loader
            image_loader = SheetImageLoader(sheet)
            for i in range(2, 3):
                file_name = sheet.cell(i, 1).value
                # get the image (put the cell you need instead of 'A1')
                label = sheet.cell(i, 3).value
                # image = image_loader.get(sheet.cell(i, 2).coordinate)
                image = sheet.cell(i, 4).value
                print(image)
                # print("**********")

                """
                from urllib.request import urlopen
                from django.core.files import File
                from django.core.files.temp import NamedTemporaryFile

                img_tmp = NamedTemporaryFile(dir='media', delete=True)
                with urlopen(image) as uo:
                    assert uo.status == 200
                    img_tmp.write(uo.read())
                    img_tmp.flush()
                img = File(img_tmp)

                question_obj = Question(is_user_credit=True)
                question_obj.save()
                question_obj.image.save(img_tmp.name, img)
                
                """
                # self.image.save(img_tmp.name, img)

                r = requests.get(image)

                if r.status_code == 200:
                    img_temp = NamedTemporaryFile(delete=True)
                    img_temp.write(r.content)
                    img_temp.flush()

                    question_obj = Question(is_user_credit=True)
                    question_obj.save()
                    question_obj.image.save(os.path.basename(image), File(img_temp), save=True)
                    question_obj.save()

                    # self.image.save(os.path.basename(self.image_url), File(img_temp), save=True)


                    print("Done", question_obj.image.url)
                print("None")
                # response = requests.get(image)
                # print(type(response.content))
                # img = Image.u(response.content)
                # img.show()
                #
                # output = BytesIO()
                # ext = img.format
                # image.save(output, format=ext, quality=95)
                # output.seek(0)
                #
                # file_obj = InMemoryUploadedFile(
                #             output,
                #             'ImageField',
                #             f"{uuid.uuid4()}.{file_name.split('.')[-1]}",
                #             f"image/{file_name.split('.')[-1]}",
                #             sys.getsizeof(output),
                #             None,
                #         )

                # is_user_credit=True for new watermarked images

                # question_obj = Question.objects.create(is_user_credit=True, image=File(open(result[0], 'rb')))

                # image = File(response.content,
                #              name=f"{uuid.uuid4()}.{file_name.split('.')[-1]}"),
                # option_filter = QuizOption.objects.filter(title__in=label.strip().split(','))
                # for opt in option_filter:
                #     question_obj.correct_option.add(opt)

            self.stdout.write('Success!!')
        # except Exception as e:
        #     self.stdout.write(f'Some error occurred while inserting quiz questions data. -- {e}')

